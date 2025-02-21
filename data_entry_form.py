import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
from tkcalendar import DateEntry
import openpyxl
import psycopg2
from datetime import date
from datetime import datetime
import logging
import re

def validate_numeric(value):
    """Validate if the input value is numeric."""
    return (value.isdigit() and len(value) <= 10 ) or value == ""  # Allow empty input or digits only

def validate_characters(value):
    """Validate if the input value is alpha numeric."""
    return value.isalpha() or value == ""  # Allow empty input or digits only


def validate_email_during_typing(value):
    """Allow typing of valid email characters and intermediate states."""
    if value == "":
        return True  # Allow empty input
    # Allow valid intermediate states
    pattern = r'^[a-zA-Z0-9._%+-]*@?[a-zA-Z0-9.-]*\.?[a-zA-Z]{0,}$'
    return re.match(pattern, value) is not None

def validate_email_final(value):
    """Check if the email is valid when completed."""
    pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
    return re.match(pattern, value) is not None

def on_focus_out(event):
    """Validate email on focus out."""
    value = email_entry.get()
    if value and not validate_email_final(value):
        messagebox.showerror("Invalid Email", "Please enter a valid email address.")
        email_entry.focus_set()  # Bring focus back to the entry widget

def validate_data(data):

    try: 
        # Check for empty fields
        for key, value in data.items():
            if not value:
                messagebox.showwarning("Validation Error", f"{key} cannot be empty.")
                logging.warning(f"Validation failed: {key} is empty.")
                return False

        # Validate email format
        email_pattern = r"[^@]+@[^@]+\.[^@]+"
        if not re.match(email_pattern, data["Email"]):
            messagebox.showwarning("Validation Error", "Invalid email format.")
            logging.warning("Validation failed: Invalid email format.")
            return False

        # Check if phone number is numeric and has at least 10 digits
        if not data["Phone"].isdigit() or len(data["Phone"]) < 10:
            messagebox.showwarning("Validation Error", "Phone number must be numeric and at least 10 digits long.")
            logging.warning("Validation failed: Phone number is not numeric or too short.")
            return False

        # Check if pin code is numeric and has at least 6 digits
        if not data["Pin Code"].isdigit() or len(data["Pin Code"]) < 6:
            messagebox.showwarning("Validation Error", "Pin code must be numeric and at least 6 digits long.")
            logging.warning("Validation failed: Pin code is not numeric or too short.")
            return False

        # Validate date of birth
        try:
            dob = datetime.strptime(data["Date of Birth"], "%m/%d/%y").date()
            today = date.today()
            age = today.year - dob.year - ((today.month, today.day) < (dob.month, dob.day))

            if dob > today:
                messagebox.showwarning("Validation Error", "Date of Birth cannot be in the future.")
                logging.warning("Validation failed: Date of Birth is in the future.")
                return False

            if age < 18:
                messagebox.showwarning("Validation Error", "Student must be at least 18 years old.")
                logging.warning("Validation failed: Student is under 18 years old.")
                return False

        except ValueError as e:
            messagebox.showwarning("Validation Error", "Invalid date format.")
            logging.warning(f"Validation failed: Invalid date format. {e}")
            return False
    
        if data["Registered"] == "Not Registered" or data["Accepted"] == "Not Accepted" :
            messagebox.showwarning("Validation Error", "You must Register and Accept terms and Condtions.")
            logging.warning("Validation failed: Course must be Registered and  Terms and Condtions must be Accept.")
            return False
        
    except Exception as err:
        messagebox.showwarning("Error", f"During validation of Field {err} error encountered")
        logging.warning(f"Validation failed for field: {err}" )
        return False

    return True
# Configure logging
logging.basicConfig(
    filename='logfile.txt',  # Log file name
    filemode='a',        # Append mode
    format='%(asctime)s - %(levelname)s - %(message)s',
    level=logging.INFO   # Log level
)


def save_data():
    # Collect data from the form
    data = {
        "Title": title_combobox.get().strip(),
        "First Name": first_name_entry.get().strip(),
        "Last Name": last_name_entry.get().strip(),
        "Gender": gender_combobox.get().strip(),
        "Date of Birth": dob_date.get().strip(),
        "Email": email_entry.get().strip(),
        "Phone": phone_entry.get().strip(),
        "Street": street_entry.get().strip(),
        "House No": house_no_entry.get().strip(),
        "Pin Code": pin_code_entry.get().strip(),
        "City": city_entry.get().strip(),
        "Course Name": course_name.get().strip(),
        "Subject 1": subject1_entry.get().strip(),
        "Subject 2": subject2_entry.get().strip(),
        "Subject 3": subject3_entry.get().strip(),
        "Registered" : registered_var.get().strip(),
        "Accepted" : accept_var.get().strip()
    }

    # Validate data
    if not validate_data(data):
        return

    # Save to Excel
    save_to_excel(data)

    # Save to PostgreSQL
    save_to_postgres(data)

def save_to_excel(data):

    try: 
    # Load or create an Excel workbook
        try:
            workbook = openpyxl.load_workbook("data.xlsx")
        except FileNotFoundError:
            workbook = openpyxl.Workbook()
            logging.info("Excel workbook not found. A new workbook has been created.")

        # Select or create a worksheet
        sheet = workbook.active
        if sheet.max_row == 1:
            # Write headers if the sheet is new
            headers = list(data.keys())
            sheet.append(headers)
            logging.info("Headers added to the new Excel sheet.")

        # Append the data
        sheet.append(list(data.values()))
        logging.info("Data appended to the Excel sheet.")

        # Save the workbook
        workbook.save("data.xlsx")
        logging.info("Excel workbook saved successfully.")
    
    except Exception as e:
        logging.error(f"Error saving to Excel: {e}")

def save_to_postgres(data):
    try:
        # Connect to PostgreSQL database
        conn = psycopg2.connect(
            dbname="student",
            user="postgres",
            password="postgres",
            host="localhost",
            port="5432"
        )
        
        conn.set_client_encoding('UTF8')
        cursor = conn.cursor()

        # Create table if it doesn't exist
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS form_data (
                id SERIAL PRIMARY KEY,
                title VARCHAR(10),
                first_name VARCHAR(50),
                last_name VARCHAR(50),
                gender VARCHAR(10),
                dob DATE,
                email VARCHAR(100),
                phone VARCHAR(15),
                street VARCHAR(100),
                house_no VARCHAR(10),
                pin_code VARCHAR(10),
                city VARCHAR(50),
                course_name VARCHAR(10),
                subject1 VARCHAR(50),
                subject2 VARCHAR(50),
                subject3 VARCHAR(50)
            )
        """)

        # Format the date to YYYY-MM-DD
        dob = datetime.strptime(data["Date of Birth"], "%m/%d/%y").strftime("%Y-%m-%d")

        # Insert data into the table
        cursor.execute("""
            INSERT INTO form_data (title, first_name, last_name, gender, dob, email, phone, street, house_no, pin_code, city, course_name, subject1, subject2, subject3)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            data["Title"], data["First Name"], data["Last Name"], data["Gender"], dob,
            data["Email"], data["Phone"], data["Street"], data["House No"], data["Pin Code"], data["City"],
            data["Course Name"], data["Subject 1"], data["Subject 2"], data["Subject 3"]
        ))

        # Commit the transaction and close the connection
        conn.commit()
        logging.info("Data saved to PostgreSQL successfully.")

    except Exception as e:
        print(f"Error saving to PostgreSQL: {e}")
        conn.rollback()
        logging.error(f"Error saving to PostgreSQL: {e}")

    finally:
        if cursor is not None:
            cursor.close()

        if conn is not None:
            conn.close()

        logging.info("Database connection closed.")

def retrieve_data():
    first_name = first_name_entry.get().strip()
    last_name = last_name_entry.get().strip()

    if not first_name or not last_name:
        messagebox.showwarning("Input Error", "First Name and Last Name must be provided to retrieve data.")
        return

    try:
        # Connect to PostgreSQL database
        conn = psycopg2.connect(
            dbname="student",
            user="postgres",
            password="postgres",
            host="localhost",
            port="5432"
        )
        
        cursor = conn.cursor()

        # Retrieve data from the table
        cursor.execute("""
            SELECT title, first_name, last_name, gender, dob, email, phone, street, house_no, pin_code, city, course_name, subject1, subject2, subject3
            FROM form_data
            WHERE first_name = %s AND last_name = %s
        """, (first_name, last_name))

        result = cursor.fetchone()

        if result:
            # Populate the form with the retrieved data
            title_combobox.set(result[0])
            first_name_entry.delete(0, tk.END)
            first_name_entry.insert(0, result[1])
            last_name_entry.delete(0, tk.END)
            last_name_entry.insert(0, result[2])
            gender_combobox.set(result[3])
            dob_date.set_date(result[4])
            email_entry.delete(0, tk.END)
            email_entry.insert(0, result[5])
            phone_entry.delete(0, tk.END)
            phone_entry.insert(0, result[6])
            street_entry.delete(0, tk.END)
            street_entry.insert(0, result[7])
            house_no_entry.delete(0, tk.END)
            house_no_entry.insert(0, result[8])
            pin_code_entry.delete(0, tk.END)
            pin_code_entry.insert(0, result[9])
            city_entry.delete(0, tk.END)
            city_entry.insert(0, result[10])
            course_name.set(result[11])
            subject1_entry.delete(0, tk.END)
            subject1_entry.insert(0, result[12])
            subject2_entry.delete(0, tk.END)
            subject2_entry.insert(0, result[13])
            subject3_entry.delete(0, tk.END)
            subject3_entry.insert(0, result[14])
            logging.info("Data retrieved and form populated successfully.")
        else:
            messagebox.showinfo("No Data Found", "No data found for the given First Name and Last Name.")
            logging.info("No data found for the given First Name and Last Name.")

    except Exception as e:
        logging.error(f"Error retrieving data from PostgreSQL: {e}")
        messagebox.showerror("Database Error", "An error occurred while retrieving data.")

    finally:
        if cursor is not None:
            cursor.close()

        if conn is not None:
            conn.close()

        logging.info("Database connection closed after retrieval.")

def update_data():
    # Collect data from the form
    data = {
        "Title": title_combobox.get().strip(),
        "First Name": first_name_entry.get().strip(),
        "Last Name": last_name_entry.get().strip(),
        "Gender": gender_combobox.get().strip(),
        "Date of Birth": dob_date.get().strip(),
        "Email": email_entry.get().strip(),
        "Phone": phone_entry.get().strip(),
        "Street": street_entry.get().strip(),
        "House No": house_no_entry.get().strip(),
        "Pin Code": pin_code_entry.get().strip(),
        "City": city_entry.get().strip(),
        "Course Name": course_name.get().strip(),
        "Subject 1": subject1_entry.get().strip(),
        "Subject 2": subject2_entry.get().strip(),
        "Subject 3": subject3_entry.get().strip(),
        "Registered" : registered_var.get().strip(),
        "Accepted" : accept_var.get().strip()
    }

    # Validate data
    if not validate_data(data):
        return

    try:
        # Connect to PostgreSQL database
        conn = psycopg2.connect(
            dbname="student",
            user="postgres",
            password="postgres",
            host="localhost",
            port="5432"
        )
        
        cursor = conn.cursor()

        # Format the date to YYYY-MM-DD
        dob = datetime.strptime(data["Date of Birth"], "%m/%d/%y").strftime("%Y-%m-%d")

        # Update data in the table
        cursor.execute("""
            UPDATE form_data
            SET title = %s, gender = %s, dob = %s, email = %s, phone = %s, street = %s, house_no = %s, pin_code = %s, city = %s, course_name = %s, subject1 = %s, subject2 = %s, subject3 = %s
             WHERE first_name = %s AND last_name = %s
        """, (
            data["Title"], data["Gender"], dob, data["Email"], data["Phone"], data["Street"],
            data["House No"], data["Pin Code"], data["City"], data["Course Name"],
            data["Subject 1"], data["Subject 2"], data["Subject 3"],
            data["First Name"], data["Last Name"]
        ))

        # Commit the transaction
        conn.commit()
        logging.info("Data updated in PostgreSQL successfully.")

        # Append the updated data to the Excel sheet
        save_to_excel(data)

    except Exception as e:
        logging.error(f"Error updating data in PostgreSQL: {e}")
        conn.rollback()
        messagebox.showerror("Database Error", "An error occurred while updating data.")

    finally:
        if cursor is not None:
            cursor.close()

        if conn is not None:
            conn.close()

        logging.info("Database connection closed after update.")

def delete_data():
    first_name = first_name_entry.get()
    last_name = last_name_entry.get()

    if not first_name or not last_name:
        messagebox.showwarning("Input Error", "First Name and Last Name must be provided to delete data.")
        return

    try:
        # Connect to PostgreSQL database
        conn = psycopg2.connect(
            dbname="student",
            user="postgres",
            password="postgres",
            host="localhost",
            port="5432"
        )
        
        cursor = conn.cursor()

        # Delete data from the table
        cursor.execute("""
            DELETE FROM form_data
            WHERE first_name = %s AND last_name = %s
        """, (first_name, last_name))

        # Check if any row was deleted
        if cursor.rowcount == 0:
            messagebox.showinfo("No Data Found", "No data found for the given First Name and Last Name.")
            logging.info("No data found for the given First Name and Last Name.")
        else:
            conn.commit()
            messagebox.showinfo("Success", "Data deleted successfully.")
            logging.info("Data deleted from PostgreSQL successfully.")

    except Exception as e:
        logging.error(f"Error deleting data from PostgreSQL: {e}")
        conn.rollback()
        messagebox.showerror("Database Error", "An error occurred while deleting data.")

    finally:
        if cursor is not None:
            cursor.close()

        if conn is not None:
            conn.close()

        logging.info("Database connection closed after deletion.")

def clear_form():
    # Clear all input fields
    title_combobox.set('')
    first_name_entry.delete(0, tk.END)
    last_name_entry.delete(0, tk.END)
    gender_combobox.set('')
    dob_date.set_date(date.today())
    email_entry.delete(0, tk.END)
    phone_entry.delete(0, tk.END)
    street_entry.delete(0, tk.END)
    house_no_entry.delete(0, tk.END)
    pin_code_entry.delete(0, tk.END)
    city_entry.delete(0, tk.END)
    course_name.set('')
    subject1_entry.delete(0, tk.END)
    subject2_entry.delete(0, tk.END)
    subject3_entry.delete(0, tk.END)

def on_mouse_wheel(event):
    # Scroll vertically
    canvas.yview_scroll(-1 * (event.delta // 120), "units")

def on_shift_mouse_wheel(event):
    # Scroll horizontally when Shift key is held
    canvas.xview_scroll(-1 * (event.delta // 120), "units")

root_window = tk.Tk()
root_window.title("Data Entry Form")
root_window.geometry("600x400+200+100")


# master frame created for complete root_window. This will cover whole area of root_window 
# it will not be visible as it is not having any border or background color.
main_frame = tk.Frame(root_window)
main_frame.pack(fill=tk.BOTH, expand=True)

# canvas is created to add scrollbar to the main_frame. It is another layer of frame which will have scrollbar.
canvas = tk.Canvas(main_frame)
canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

# vertical scrollbar is created and attached to the canvas. It will scroll the canvas vertically.
v_scrollbar = ttk.Scrollbar(main_frame, orient=tk.VERTICAL, command=canvas.yview)
v_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

# horizontal scrollbar is created and attached to the canvas. It will scroll the canvas horizontally.
h_scrollbar = ttk.Scrollbar(root_window, orient=tk.HORIZONTAL, command=canvas.xview)
h_scrollbar.pack(side=tk.BOTTOM, fill=tk.X)

# configure the canvas to use the scrollbar. It will scroll the canvas vertically and horizontally.
canvas.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)
# bind the canvas to the scrollbar. It will scroll the canvas vertically and horizontally.
canvas.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))

# Bind mouse scroll events
canvas.bind_all("<MouseWheel>", on_mouse_wheel)
canvas.bind_all("<Shift-MouseWheel>", on_shift_mouse_wheel)

# scrollbar_frame is created to add all the widgets. This frame is added to the canvas.
scrollbar_frame = tk.Frame(canvas)
canvas.create_window((0,0), window=scrollbar_frame, anchor="nw")

# frame1 is created to add personal information widgets. This frame is added to the scrollbar_frame.
frame1 = tk.Frame(scrollbar_frame, bd=6, relief=tk.RAISED, bg="lightcyan")  # relief : flat, groove, raised, ridge, solid, or sunken
frame1.pack(padx=10, pady=10)

# personal_info_frame is created to add personal information widgets. This frame is added to the frame1.
personal_info_frame = tk.LabelFrame(frame1,text="Personal Information", bg="lightgrey")
personal_info_frame.grid(row=0,column=0) #sticky="nsew",padx=20,pady=10)

# personal information widgets are added to the personal_info_frame.
title_label = tk.Label(personal_info_frame, text="Title")
title_label.grid(row=0,column=0)

title_combobox = ttk.Combobox(personal_info_frame, values=["Mr","Mrs","Miss"])
title_combobox.grid(row=1,column=0)

first_name_lable = tk.Label(personal_info_frame, text="First Name")
first_name_lable.grid(row=0,column=1) 

first_name_entry = tk.Entry(personal_info_frame)
first_name_entry.grid(row=1,column=1)

last_name_lable = tk.Label(personal_info_frame, text="Last Name")
last_name_lable.grid(row=0,column=2) 

last_name_entry = tk.Entry(personal_info_frame)
last_name_entry.grid(row=1,column=2)

gender_label = tk.Label(personal_info_frame, text="Gender")
gender_label.grid(row=2,column=0)

#Combobox is used to create dropdown list.
gender_combobox = ttk.Combobox(personal_info_frame, values=["Male", "Female"])
gender_combobox.grid(row=3,column=0)

dob_label = tk.Label(personal_info_frame, text="Date of Birth")
dob_label.grid(row=2,column=1)

#DateEntry is used to create date picker.
dob_date = DateEntry(personal_info_frame, selectmode="day")
dob_date.set_date(date.today())
dob_date.grid(row=3,column=1)

# padx and pady are used to add padding to the widgets.
for widget in personal_info_frame.winfo_children():
    widget.grid(padx=10,pady=10)

contact_info_frame = tk.LabelFrame(frame1,text="Contact Information", bg="lightgrey")
contact_info_frame.grid(row=1, column=0, padx=10, pady=10)

email_label = tk.Label(contact_info_frame, text="Email")
email_label.grid(row=0, column=0)

vcmd_email = root_window.register(validate_email_during_typing)

email_entry = tk.Entry(contact_info_frame, validate="key", validatecommand=(vcmd_email, "%P"))
email_entry.grid(row=1,column=0)

# Bind the focus out event for final validation
email_entry.bind("<FocusOut>", on_focus_out)

phone_label = tk.Label(contact_info_frame, text="Phone")
phone_label.grid(row=0,column=1)

vcmd_phone =  root_window.register(validate_numeric)

phone_entry = tk.Entry(contact_info_frame, validate="key", validatecommand=(vcmd_phone, "%P"))
phone_entry.grid(row=1,column=1)

for widget in contact_info_frame.winfo_children():
    widget.grid(padx=10,pady=10)

# frame2 is created to add address widgets. This frame is added to the scrollbar_frame.
frame2 = tk.Frame(scrollbar_frame, bd=6, relief=tk.SUNKEN, bg="lightgreen")  # relief : flat, groove, raised, ridge, solid, or sunken
frame2.pack(padx=10, pady=10)

# address_frame is created to add address widgets. This frame is added to the frame2.
address_frame = tk.LabelFrame(frame2,text="Address", bg="lightgrey", padx=10, pady=10)
address_frame.grid(row=0,column=0) #sticky="nsew",padx=20,pady=10)

# address widgets are added to the address_frame.
street_label = tk.Label(address_frame, text="Street")
street_label.grid(row=0,column=0)

street_entry = tk.Entry(address_frame)
street_entry.grid(row=1,column=0)

house_no_label = tk.Label(address_frame, text="House No")
house_no_label.grid(row=0,column=1)

vcmd_house_no =  root_window.register(validate_numeric)
house_no_entry = tk.Entry(address_frame, validate="key", validatecommand=(vcmd_house_no, "%P"))
house_no_entry.grid(row=1,column=1)

pin_code_label = tk.Label(address_frame, text="Pin Code")
pin_code_label.grid(row=2,column=0)

vcmd_pin_code =  root_window.register(validate_numeric)
pin_code_entry = tk.Entry(address_frame, validate="key", validatecommand=(vcmd_pin_code, "%P"))
pin_code_entry.grid(row=3,column=0)

city_label = tk.Label(address_frame, text="City")
city_label.grid(row=2,column=1)

city_entry = tk.Entry(address_frame)
city_entry.grid(row=3,column=1)

for widget in address_frame.winfo_children():
    widget.grid(padx=10,pady=10)

# frame3 is created to add course widgets. This frame is added to the scrollbar_frame.
frame3 = tk.Frame(scrollbar_frame, bd=6, relief=tk.RIDGE, bg="gold")  # relief : flat, groove, raised, ridge, solid, or sunken
frame3.pack(padx=10, pady=10)

# course_frame is created to add course widgets. This frame is added to the frame3.
course_frame = tk.LabelFrame(frame3,text="Courses", bg="lightgrey", padx=10, pady=10)
course_frame.grid(row=0,column=0) #sticky="nsew",padx=20,pady=10)

course_label = tk.Label(course_frame, text="Course Name")
course_label.grid(row=0,column=0)

course_name = ttk.Combobox(course_frame, values=["FE","SE","TE","BE"])
course_name.grid(row=1,column=0)

subject1_label = tk.Label(course_frame, text="Subject 1")
subject1_label.grid(row=0,column=1)

subject1_entry = tk.Entry(course_frame)
subject1_entry.grid(row=1,column=1)

subject2_label = tk.Label(course_frame, text="Subject 2")
subject2_label.grid(row=0,column=2)

subject2_entry = tk.Entry(course_frame)
subject2_entry.grid(row=1,column=2)

subject3_label = tk.Label(course_frame, text="Subject 3")
subject3_label.grid(row=0,column=3)

subject3_entry = tk.Entry(course_frame)
subject3_entry.grid(row=1,column=3)

for widget in course_frame.winfo_children():
    widget.grid(padx=10,pady=10)

# frame4 is created to add action widgets. This frame is added to the scrollbar_frame.
frame4 = tk.Frame(scrollbar_frame, bd=6, relief=tk.RIDGE, bg="fuchsia")  # relief : flat, groove, raised, ridge, solid, or sunken
frame4.pack(padx=10, pady=10)

# action_frame is created to add action widgets. This frame is added to the frame4.
terms_frame = tk.LabelFrame(frame4,text="Terms and Conditions", bg="lightgrey", padx=10, pady=10)
terms_frame.grid(row=0,column=0) #sticky="nsew",padx=20,pady=10)

registered_var = tk.StringVar(value="Not Registered")
registered_check = tk.Checkbutton(terms_frame, text= "I would like to register for above course.",
                                  variable=registered_var, onvalue="Registered", offvalue="Not Registered")
registered_check.grid(row=0, column=0)

accept_var = tk.StringVar(value="Not Accepted")
terms_check = tk.Checkbutton(terms_frame, text= "I accept the terms and conditions.",
                                  variable=accept_var, onvalue="Accepted", offvalue="Not Accepted")
terms_check.grid(row=1, column=0)

for widget in terms_frame.winfo_children():
    widget.grid(padx=5, pady=5)


# frame4 is created to add action widgets. This frame is added to the scrollbar_frame.
frame5 = tk.Frame(scrollbar_frame, bd=6, relief=tk.RIDGE, bg="silver")  # relief : flat, groove, raised, ridge, solid, or sunken
frame5.pack(padx=10, pady=10)

# action_frame is created to add action widgets. This frame is added to the frame4.
action_frame = tk.LabelFrame(frame5,text="Press Button", bg="lightgrey", padx=10, pady=10)
action_frame.grid(row=0,column=0) #sticky="nsew",padx=20,pady=10)


save_button = tk.Button(action_frame, text="Save", command=lambda :save_data())
save_button.grid(row=0,column=0)

# Update button
update_button = tk.Button(action_frame, text="Update", command=update_data)
update_button.grid(row=0, column=10)

# Delete button
delete_button = tk.Button(action_frame, text="Delete", command=delete_data)
delete_button.grid(row=0, column=15)

clear_button = tk.Button(action_frame, text="Clear", command=lambda :clear_form())
clear_button.grid(row=0,column=20)

# Retrieve button
retrieve_button = tk.Button(action_frame, text="Retrieve", command=retrieve_data)
retrieve_button.grid(row=0, column=25)

close_button = tk.Button(action_frame, text="Close", command=root_window.quit)
close_button.grid(row=0,column=30)

for widget in action_frame.winfo_children():
    widget.grid(padx=20,pady=20)

root_window.mainloop()